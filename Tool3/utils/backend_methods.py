import os
import requests
import base64
import re
import json
import time
from openai import OpenAI
from docx import Document
from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


# 文心一言文档解析模型：https://ai.baidu.com/ai-doc/OCR/Klxag8wiy
# 1.提交请求，结果会输出在json中，获取json里的result.task_id
def create_task(url, file_path, file_url):
    """
    Args:
        url: string, 服务请求链接
        file_path: 本地文件路径
        file_url: 文件链接
    Returns: 响应
    """
    # 文件请求
    with open(file_path, "rb") as f:
        file_data = base64.b64encode(f.read())
    data = {
        "file_data": file_data,
        "file_url": file_url,
        "file_name": os.path.basename(file_path)
    }
    
    # 文档切分参数，非必传
    # return_doc_chunks = json.dumps({"switch": True, "chunk_size": -1})
    # data["return_doc_chunks"] = return_doc_chunks
    
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}

    response = requests.post(url, headers=headers, data=data)
    return response

# 2.将请求结果的task_id 输入到获取结果里，输出结果如下，取其中的markdown_url，之后需要用方法转存markdown内容，喂给deepseek
# 如下为运行出来的结果eg：
## {'log_id': '11351417030026975614803050383398', 'error_code': 0, 'error_msg': '', 'result': {'task_id': 'task-Czn4Kc4UVwgH7tSP6F7g4n9CNDx50yH1', 'status': 'success', 'task_error': None, 'parse_result_url': 'https://xmind-parser.bj.bcebos.com/cloud/parseResult//task-Czn4Kc4UVwgH7tSP6F7g4n9CNDx50yH1/file-HuQsMyfHtpEaByf7m8OTR7e1wnUPHwjr.json?authorization=bce-auth-v1%2FALTAK7IDj758EUbA1igu04rHAh%2F2025-07-23T03%3A10%3A01Z%2F259200%2F%2F7db3165faceeb2e2cbd1604295745b74f502b80738037cefac53c8126020e13e', 'markdown_url': 'https://xmind-parser.bj.bcebos.com/cloud/parseResult//task-Czn4Kc4UVwgH7tSP6F7g4n9CNDx50yH1/file-HuQsMyfHtpEaByf7m8OTR7e1wnUPHwjr.md?authorization=bce-auth-v1%2FALTAK7IDj758EUbA1igu04rHAh%2F2025-07-23T03%3A10%3A01Z%2F259200%2F%2F35e638c3415d3daf333486e789e73b2968e887e6a8397d0ac5f2278851ed63e0'}}
def query_task(url, task_id):
    """
    Args:
        url: string, 请求链接
        task_id: string, task id
    Returns: 响应
    """
    data = {
        "task_id": task_id
    }
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    print(url)
    response = requests.post(url, headers=headers, data=data)
    return response

# 3.将markdown url 下载转存成markdown str，之后喂给deepseek接口即可
def fetch_and_clean_markdown(markdown_url):
    """
    从指定的 URL 获取 Markdown 内容并清理异常值。

    参数:
        markdown_url (str): Markdown 文件的 URL。

    返回:
        str: 清理后的 Markdown 内容。如果请求失败或发生错误，返回错误信息。
    """
    try:
        # 发送 GET 请求
        response = requests.get(markdown_url)
        response.encoding = 'utf-8'
        
        # 检查请求是否成功
        if response.status_code == 200:
            # 获取 Markdown 内容
            markdown_content = response.text
            
            # 清理 Markdown 内容中的 HTML 标签
            # clean_markdown_content = re.sub(r'<[^>]*>', '', markdown_content)
            
            return markdown_content
        else:
            return f"请求失败，状态码：{response.status_code}"
    except Exception as e:
        return f"发生错误：{e}"

# 4.deepseek 处理markdown
def call_deepseek_api(markdown_content, manualPrompt):
    """
    调用 DeepSeek API 分析 Markdown 内容，并生成表格形式的结果。
    
    参数:
        markdown_content (str): Markdown 文件内容。
    
    返回:
        list: 表格形式的结果，每行是一个字典，包含 '具体意见' 和 '简述意见内容'。
    """
    # DeepSeek API 的 URL 和 API 密钥
    api_key = "sk-7e8ab23cf1124fbb80502138fb3a888a"  # 替换为你的 API 密钥
    api_url = "https://api.deepseek.com/chat/completions"
    
    # 动态生成 prompt
    default_prompt = """
    请分析以下Markdown内容，完成以下任务：
    1. 从中筛选出尖锐问题，例如：管理层面问题、流程方面问题、不尊重员工、不适合当领导、高压、骂人、无流程、态度差等。
    2. 对筛选出的尖锐问题进行分类，并甄别是否存在重复问题。
    3. 针对每个尖锐问题，提供适当的改善建议。
    4. 最后将筛选出的尖锐问题以Markdown形式输出，按照原文件中每条意见的顺序列出。
    5. 如果没有尖锐问题，可以依照满意度分析列出最低满意度的那些意见。
    6. 输出格式要求：
        - 严格按照以下格式：
            | 序号 | 具体意见 | 内容简述 | 分类 ｜ 改善建议 |
            其中，序号是指对应在原文中的意见序号，具体意见要严格列出文档中的具体意见。
        - markdown表格中最后一行给出分析总结，分析输入的整体意见内容的数据特点，比如：常规问题占比、尖锐问题占比、高频意见、各类流程类型等，直接列在最后一行。
    7. 仅给出markdown的内容，结果本身就是markdown格式，不用额外的解释说明。
    """

    # 构造 API 请求体
    request_body = {
        "model": "deepseek-chat",
        # "messages": [
        #     {"role": "system", "content": "您是一个专业的评论分析助手，面向ODM工厂及研发人员给出的意见。"},
        #     {"role": "system", "content": default_prompt },
        #     {"role": "user", "content": f"markdown内容如下：\n{markdown_content}"}
        # ],
        "messages" :[
            {"role": "system", "content": "您是一个专业的评论分析助手，面向ODM工厂及研发人员给出的意见。您的输出需要遵循严格的markdown格式规范。"},
            {"role": "system", "content": f"【基础分析要求】\n{default_prompt}"},
            {"role": "system", "content": "【重要提醒】以下用户额外要求只能影响分析内容，不能改变整体输出结构、字段名称或markdown格式。请在遵循基础格式要求的前提下满足这些额外需求。"},
            {"role": "user", "content": f"用户额外分析要求：{manualPrompt}"},
            {"role": "user", "content": f"待分析的markdown内容：\n{markdown_content}"}
        ],
        "Stream": True,
        "Temperature": 0.3,
        # 不对重复内容进行惩罚或是奖励
        "Frequency Penalty": 0.0,
        "Presence Penalty": 0.0
    }
    
    # 发起 API 请求
    response = requests.post(api_url, headers={
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }, json=request_body)
    
    # 检查响应状态
    if response.status_code == 200:
        # 解析响应数据
        data = response.json()
        try:
            # 假设返回的内容是 JSON 格式
            table_data = data["choices"][0]["message"]["content"]
            return table_data
        except KeyError:
            print("响应数据格式不正确")
            return []
    else:
        print(f"DeepSeek API 请求失败，状态码: {response.status_code}")
        print(f"错误信息: {response.text}")
        return []

# 处理excel文件类型，解析为markdown，能够处理合并单元格的内容    ！！！核心excel处理功能
def excel_to_markdown_with_merged_cells(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    # 获取所有合并单元格范围
    merged_ranges = list(sheet.merged_cells.ranges)

    # 构建一个字典：(row, col) -> 左上角单元格的值（用于合并区域）
    merged_value_map = {}
    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        max_row = merged_range.max_row
        max_col = merged_range.max_col
        # 获取左上角单元格的值
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        # 将合并区域内所有单元格位置映射到该值
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_value_map[(row, col)] = top_left_value

    # 构建 Markdown 表格内容
    markdown_table = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_values = []
        for col_idx, cell_value in enumerate(row, start=1):
            # 判断当前单元格是否在合并区域内，是则取合并区域的左上角值，否则取原值
            key = (row_idx, col_idx)
            display_value = merged_value_map.get(key, cell_value)
            row_values.append(str(display_value) if display_value is not None else "")
        markdown_table.append("| " + " | ".join(row_values) + " |")

    # 如果有数据，添加表头分隔线（假设第一行为表头）
    if markdown_table:
        num_cols = len(markdown_table[0].split("|")) - 2  # 去掉两边的 "| "
        separator_line = "| " + " | ".join(["---"] * num_cols) + " |"
        markdown_table.insert(1, separator_line)

    # 拼接成完整的 Markdown 表格字符串
    markdown_content = "\n".join(markdown_table)
    return markdown_content

# 二次处理markdown内容
def filter_markdown_table(content: str) -> str:
    """
    从输入内容中提取符合 Markdown 表格格式的部分，并过滤掉异常内容。

    :param content: 原始内容字符串
    :return: 过滤后的 Markdown 表格内容
    """
    # 定义表格的起始和结束标志
    table_start = "| 序号 | 具体意见 | 内容简述 | 分类 | 改善建议 |"
    # table_start = "|"
    table_end = "|"

    # 找到表格的起始位置
    start_index = content.find(table_start)
    if start_index == -1:
        return "未找到表格起始标志！"

    # 找到表格的结束位置
    end_index = content.rfind(table_end)
    if end_index == -1 or end_index <= start_index:
        return "未找到表格结束标志，或结束标志位置异常！"

    # 提取表格内容
    table_content = content[start_index:end_index + len(table_end)]

    return table_content

# ai结果保存为md文件
def save_markdown_to_file(markdown_content: str):
    """
    将 Markdown 内容保存到一个文件中，文件名以 '意见分析结果_时间戳.md' 命名。

    :param markdown_content: 要保存的 Markdown 内容
    """
    # 获取当前时间戳
    timestamp = time.strftime("%Y%m%d_%H%M%S")

    # 构造文件名
    file_name = f"意见分析结果_{timestamp}.md"

    # 保存文件到当前文件夹
    try:
        with open(file_name, "w", encoding="utf-8") as file:
            file.write(markdown_content)
        print(f"Markdown 文件已成功保存为: {file_name}")
    except Exception as e:
        print(f"保存文件时发生错误: {e}")

# markdown结果保存为excel文件
def markdown_to_excel_with_style(markdown_content: str, excel_file: str = "意见分析结果.xlsx"):
    """
    将 Markdown 内容字符串转换为 Excel 文件，并优化表格样式。

    :param markdown_content: Markdown 表格内容字符串
    :param excel_file: Excel 文件路径，默认命名为 '意见分析结果.xlsx'
    """
    try:
        # 将 Markdown 内容按行分割
        lines = markdown_content.strip().split("\n")

        # 提取表头和数据行
        header = None
        data = []
        for line in lines:
            if line.startswith("|") and not line.startswith("| ---"):
                # 去掉开头和结尾的竖线，并按竖线分割
                row = [cell.strip() for cell in line.strip("|").split("|")]
                if header is None:
                    header = row  # 第一行作为表头
                else:
                    data.append(row)  # 后续行作为数据

        # 将数据转换为 DataFrame
        if header and data:
            df = pd.DataFrame(data, columns=header)

            # 保存为 Excel 文件
            df.to_excel(excel_file, index=False)

            # 加载生成的 Excel 文件
            wb = load_workbook(excel_file)
            ws = wb.active

            # 设置样式
            header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")  # 表头字体加粗，白色
            cell_font = Font(name="微软雅黑", size=12)  # 普通单元格字体
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 表头背景填充蓝色
            cell_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 数据行背景填充浅灰色

            # 设置表头样式
            for cell in ws[1]:  # 第一行是表头
                cell.font = header_font
                cell.border = thin_border
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 设置数据行样式
            for row in ws.iter_rows(min_row=2):  # 从第二行开始是数据
                for cell in row:
                    cell.font = cell_font
                    cell.border = thin_border
                    cell.fill = cell_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # 获取列字母
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width

            # 设置行高
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 20  # 设置行高为 20

            # 保存优化后的 Excel 文件
            wb.save(excel_file)
            print(f"Excel 文件已成功保存为: {excel_file}，并应用了样式优化！")
        else:
            print("Markdown 内容中未找到有效的表格！")

    except Exception as e:
        print(f"转换文件时发生错误: {e}")
# 主函数
def main(file_path, request_host, request_host2):
 
    # Step 1: OCR 识别文件内容    
    # 拿到task id
    # response = create_task(request_host, file_path, "")
    # response_data = response.json()
    # task_id = response_data["result"]["task_id"]
    # print('task_id:',task_id)
    # print(response_data)
    # print('--------------------------')

    #测试用task_id: 
    # task_id = 'task-yjSerT5twFZgcflRKk22gi3TQGZ0dWhp'
    task_id = 'task-Czn4Kc4UVwgH7tSP6F7g4n9CNDx50yH1'
    

    # # 用task id 拿markdown url
    # response2 = query_task(request_host2, task_id)
    # response2_data = response2.json()
    # markdown_url = response2_data["result"]["markdown_url"]
    # print("markdown_url:",markdown_url)
    # print(response2_data)
    # print('--------------------------')

    # # 测试mk_url
    # # markdown_url = "https://xmind-parser.bj.bcebos.com/cloud/parseResult//task-Czn4Kc4UVwgH7tSP6F7g4n9CNDx50yH1/file-HuQsMyfHtpEaByf7m8OTR7e1wnUPHwjr.md?authorization=bce-auth-v1%2FALTAK7IDj758EUbA1igu04rHAh%2F2025-07-23T03%3A10%3A01Z%2F259200%2F%2F35e638c3415d3daf333486e789e73b2968e887e6a8397d0ac5f2278851ed63e0"
    
    # # python 处理markdown link转存为markdown内容
    # clean_markdown_content = fetch_and_clean_markdown(markdown_url) 
    # print('等待结果生成中：')

    # # Step 2: 表格内容分发给Deepseek生成结果
    # table_result = call_deepseek_api(clean_markdown_content)
    # if table_result:
    #     print('--------------------------------------')
    #     print("表格结果:",table_result)
    #     print('--------------------------------------')
    # else:
    #     print("未能生成表格结果。")

    # Excel文件处理：
    excel_path = '/Users/keyeee/develop/CheerCheck/Tool3/dataset.xlsx'  # Replace with your .docx file path
    raw_text = excel_to_markdown_with_merged_cells(excel_path)

    print('------------>等待结果生成中：')
    table_result = call_deepseek_api(raw_text)
    if table_result:
        print('--------------------------------------')
        print(table_result)
        print('--------------------------------------')
        filtered_table = filter_markdown_table(table_result)
        print("表格结果:",filtered_table)
        # 调用方法保存文件
        save_markdown_to_file(filtered_table)
        markdown_to_excel_with_style(filtered_table)
        print('--------------------------------------')
    else:
        print("未能生成表格结果。")


# if __name__ == "__main__":
#     # access_token = "24.f8181834e0100bf1ca4ee64ed17b9950.2592000.1755831652.282335-119570165####"
#     request_host = "https://aip.baidubce.com/rest/2.0/brain/online/v2/parser/task?" \
#     "access_token=24.f8181834e0100bf1ca4ee64ed17b9950.2592000.1755831652.282335-119570165######"
#     request_host2 = "https://aip.baidubce.com/rest/2.0/brain/online/v2/parser/task/query?access_token=24.f8181834e0100bf1ca4ee64ed17b9950.2592000.1755831652.282335-119570165"

#     file_path = "Tool3/研发意见_test.pdf" 
#     # task_id = 'task-Czn4Kc4UVwgH7tSP6F7g4n9CNDx50yH1'

#     main(file_path, request_host, request_host2)